#!/usr/bin/env python
# coding=utf-8
'''
作    者 : 北极星光 light22@126.com
创建时间 : 2024-06-02 00:21:43
最后修改 : 2024-11-04 16:07:57
修 改 者 : 北极星光
'''

import os
from datetime import datetime
from docx import Document
from openpyxl import load_workbook
from tools import *


def auto_settle_protocol():
    # 读取config.xlsx文件
    workbook = load_workbook('config.xlsx', data_only=True)
    sheet = workbook['结算协议']

    # 遍历sheet
    for i in range(2, sheet.max_row + 1):
        protocol_name = sheet[f'A{i}'].value  # 框架协议名称 <protocol_name>
        contract_number = sheet[f'B{i}'].value  # 框架协议编号 <contract_number>
        contract_sign_date = sheet[f'C{i}'].value.strftime('%Y年%m月') if type(sheet[f'C{i}'].value) == datetime else excel_date_format(sheet[f'C{i}'].value).strftime('%Y年%m月')  # 框架协议签订年月 <contract_sign_date>
        project_name = sheet[f'D{i}'].value  # 工程名称 <project_name>
        project_address = sheet[f'E{i}'].value  # 工程地址 <project_address>
        party_a_name = sheet[f'F{i}'].value  # 甲方名称 <party_a_name>
        party_b_name = sheet[f'G{i}'].value  # 乙方名称 <party_b_name>
        completion_date = sheet[f'H{i}'].value.strftime('%Y年%m月%d日') if type(sheet[f'H{i}'].value) == datetime else excel_date_format(sheet[f'I{i}'].value).strftime('%Y年%m月%d日')  # 竣工日期 <completion_date>
        warranty_period = sheet[f'J{i}'].value  # 质保期
        settlement_sign_date = sheet[f'K{i}'].value.strftime('%Y年%m月') if type(sheet[f'K{i}'].value) == datetime else excel_date_format(sheet[f'K{i}'].value).strftime('%Y年%m月')  # 结算协议签订年月 <settlement_sign_date>
        settlement_amount = round(sheet[f'L{i}'].value, 2)  # 结算金额 <settlement_amount>
        settlement_amount_capital = num_to_capital(settlement_amount)  # 结算金额大写 <settlement_amount_capital>
        tax_rate = sheet[f'M{i}'].value  # 税率
        tax_rate_percent = f'{tax_rate * 100:.0f}%'  # 税率百分比 <tax_rate_percent>
        settlement_amount_untax = round(settlement_amount / (1 + tax_rate),
                                        2)  # 结算不含税金额 <settlement_amount_untax>
        settlement_amount_tax = round(settlement_amount -
                                      settlement_amount_untax, 2)  # 结算税金 <settlement_amount_tax>
        paid_amount = round(sheet[f'N{i}'].value, 2)  # 已支付金额 <paid_amount>
        paid_amount_capital = num_to_capital(paid_amount)  # 已支付金额大写 <paid_amount_capital>
        balance_amount = round(settlement_amount - paid_amount, 2)  # 剩余金额 <balance_amount>
        balance_amount_capital = num_to_capital(balance_amount)  # 剩余金额大写 <balance_amount_capital>
        settlement_amount_untax_balance = round(balance_amount /
                                                (1 + tax_rate), 2)  #剩余不含税金额 <settlement_amount_untax_balance>
        settlement_amount_tax_balance = round(
            balance_amount - settlement_amount_untax_balance, 2)  # 剩余税金 <settlement_amount_tax_balance>
        payment_ratio = sheet[f'O{i}'].value  # 结算后支付比例
        payment_ratio_percent = f'{payment_ratio * 100:.0f}%'  # 结算后支付比例百分比 <payment_ratio_percent>
        payment_amount = round(settlement_amount * payment_ratio,
                               2)  # 结算后应支付金额 <payment_amount>
        payment_amount_capital = num_to_capital(payment_amount)  # 结算后应支付金额大写 <payment_amount_capital>
        payment_amount_this = round(payment_amount - paid_amount, 2)  # 本次付款金额 <payment_amount_this>
        payment_amount_this_capital = num_to_capital(
            payment_amount_this)  # 本次付款金额大写 <payment_amount_this_capital>
        warranty_ratio = sheet[f'P{i}'].value  # 质保金比例
        warranty_ratio_percent = f'{warranty_ratio * 100:.0f}%'  # 质保金比例百分比 <warranty_ratio_percent>
        warranty_amount = round(settlement_amount * warranty_ratio, 2)  # 质保金金额 <warranty_amount>
        warranty_amount_capital = num_to_capital(warranty_amount)  # 质保金金额大写 <warranty_amount_capital>

        # 创建工程文件夹
        path = f'output/{project_name}工程结算协议'
        os.makedirs(path, exist_ok=True)

        # 生成结算协议
        document = Document('data/结算协议模板/结算协议-模板.docx')
        # 替换模板中的变量
        replace_variables(
            document,
            protocol_name=protocol_name,
            contract_number=contract_number,
            contract_sign_date=contract_sign_date,
            project_name=project_name,
            party_a_name=party_a_name,
            party_b_name=party_b_name,
            completion_date=completion_date,
            warranty_period=warranty_period,
            settlement_sign_date=settlement_sign_date,
            settlement_amount=settlement_amount,
            settlement_amount_capital=settlement_amount_capital,
            tax_rate=tax_rate,
            tax_rate_percent=tax_rate_percent,
            settlement_amount_untax=settlement_amount_untax,
            settlement_amount_tax=settlement_amount_tax,
            paid_amount=paid_amount,
            paid_amount_capital=paid_amount_capital,
            balance_amount=balance_amount,
            balance_amount_capital=balance_amount_capital,
            settlement_amount_untax_balance=settlement_amount_untax_balance,
            settlement_amount_tax_balance=settlement_amount_tax_balance,
            payment_ratio=payment_ratio,
            payment_ratio_percent=payment_ratio_percent,
            payment_amount=payment_amount,
            payment_amount_capital=payment_amount_capital,
            payment_amount_this=payment_amount_this,
            payment_amount_this_capital=payment_amount_this_capital,
            warranty_ratio=warranty_ratio,
            warranty_ratio_percent=warranty_ratio_percent,
            warranty_amount=warranty_amount,
            warranty_amount_capital=warranty_amount_capital)

        # 保存文件
        document.save(f'{path}/{project_name}工程结算协议.docx')
        print(f'生成{project_name}工程结算协议成功！')

        # 生成质量保修书
        document = Document('data/结算协议模板/质量保修书-模板.docx')

        # 替换模板中的变量
        replace_variables(document,
                          protocol_name=protocol_name,
                          project_name=project_name,
                          project_address=project_address,
                          party_a_name=party_a_name,
                          party_b_name=party_b_name,
                          completion_date=completion_date,
                          tax_rate_percent=tax_rate_percent,
                          warranty_ratio_percent=warranty_ratio_percent,
                          warranty_amount=warranty_amount,
                          warranty_amount_capital=warranty_amount_capital)

        # 保存文件
        document.save(f'{path}/{project_name}工程质量保修书.docx')
        print(f'生成{project_name}工程质量保修书成功！')
